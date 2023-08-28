from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from django.utils import formats
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from openpyxl.utils import get_column_letter
from django.db.models import Sum

from .models import Medicaments, Transactions, Travailleurs, Fournisseurs
from .forms import MedicamentsForm, TravailleursForm, FournisseursForm

# Create your views here.

class HomeMedocView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Medicaments 
    template_name = "stocks_list.html"
    context_object_name = "medocs"

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        medicaments = Medicaments.objects.all()
        paginator = Paginator(medicaments, 10) #chaque page affiche 10medicaments
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['medocs'] = page_obj
        return context
    
    #On fusionne les dispatch des differentes classes (le dispatch de loginrequiredmixin, celui du listview et userpassestestmixin)
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        date = request.POST.get('date')
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')
        nom_medoc = request.POST.get('nom_medoc')

        medocs = Medicaments.objects.all()
    
        if date:
            medocs = Medicaments.search_by_date(date)
        if mois:
            medocs = Medicaments.search_by_month(mois)
        if annee:
            medocs = Medicaments.search_by_year(annee)
        if nom_medoc:
            medocs = Medicaments.search_by_name(nom_medoc)

        context = {
            'medocs': medocs
        }

        return render(request, 'stocks_list.html', context)
    

class CreateMedocView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Medicaments
    template_name = "stocks_create.html"
    form_class = MedicamentsForm
    success_url = reverse_lazy('stocks:list_medocs')

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()

    def form_invalid(self, form):
        # Renvoie les erreurs de validation sous forme de réponse JSON
        return JsonResponse({'errors': form.errors})
    
    def form_valid(self, form):
        # Enregistrer le formulaire si valide
        self.object = form.save()
        # Renvoyer une réponse JSON avec succès
        return JsonResponse({'success': True}) 

class UpdateMedocView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Medicaments
    template_name = "stocks_update.html"
    form_class = MedicamentsForm
    success_url = reverse_lazy('stocks:list_medocs')

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()    


class TransactionView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Medicaments 
    context_object_name = "medocs"
    template_name = "retrait.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        medicaments = Medicaments.objects.all()
        paginator = Paginator(medicaments, 10) #chaque page affiche 10medicaments
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['medocs'] = page_obj
        return context
    
    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func    
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        nom_medoc = request.POST.get('nom_medoc')

        medocs = Medicaments.objects.all()
    
        if nom_medoc:
            medocs = Medicaments.search_by_name(nom_medoc)

        context = {
            'medocs': medocs
        }

        return render(request, 'retrait.html', context)
    

class RechercheView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Medicaments 
    context_object_name = "medocs"
    template_name = "transaction/transaction_create.html"

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        medicaments = Medicaments.objects.all()
        paginator = Paginator(medicaments, 10) #chaque page affiche 10medicaments
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['medocs'] = page_obj
        return context
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        nom_medoc = request.POST.get('nom_medoc')

        medocs = Medicaments.objects.all()
    
        if nom_medoc:
            medocs = Medicaments.search_by_name(nom_medoc)

        context = {
            'medocs': medocs
        }

        return render(request, 'transaction/transaction_create.html', context)
    

class HomeTransactionView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Transactions
    template_name = "transaction/transaction_list.html"
    context_object_name = "transactions"

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        transactions = Transactions.objects.all()
        paginator = Paginator(transactions, 10) #chaque page affiche 10transactions
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['transactions'] = page_obj
        return context
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        date = request.POST.get('date')
        mois = request.POST.get('mois')
        annee = request.POST.get('annee')
        medicaments = request.POST.get('medicaments')
        categorie = request.POST.get('categorie')

        transactions = Transactions.objects.all()
    
        if date:
            transactions = Transactions.search_by_date(date)
        if mois:
            transactions = Transactions.search_by_month(mois)
        if annee:
            transactions = Transactions.search_by_year(annee)
        if medicaments:
            transactions = Transactions.search_by_name(medicaments)
        if categorie:
            transactions = transactions.filter(type_transaction=categorie)


        context = {
            'transactions': transactions
        }

        return render(request, 'transaction/transaction_list.html', context)


@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def achat(request, slug):
    medicament = get_object_or_404(Medicaments, slug= slug)
    fournisseurs = Fournisseurs.objects.all()
    
    if request.method == 'POST':
        quantite = int(request.POST['quantite'])
        qtite_plaq = int(request.POST['quantite_plaq'])
        fournisseur_id = int(request.POST['fournisseur'])
    
        # Vérifiez si un fournisseur a été sélectionné
        if fournisseur_id:
            fournisseur = Fournisseurs.objects.get(id=fournisseur_id)
                
            medicament.quantité_stock += quantite
            medicament.quantité_detail += qtite_plaq    
            medicament.quantité_stock += qtite_plaq // medicament.plaquette_par_boite
            medicament.quantité_detail += quantite * medicament.plaquette_par_boite
            medicament.save()
         
            transaction = Transactions(medicaments=medicament,type_transaction=Transactions.ACHAT, quantite=quantite, quantite_plaq=qtite_plaq, fournisseur=fournisseur, user=request.user)
            transaction.save(request=request)
        
            return redirect('stocks:list_transaction')
        else:
            messages.error(request, "Veuillez sélectionner un fournisseur.")
        
    return render(request, 'transaction/ajouter_medoc.html', {'medicament': medicament, 'fournisseurs':fournisseurs})

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def vente(request, slug):
    medicament = get_object_or_404(Medicaments, slug= slug)
    travailleurs = Travailleurs.objects.all()
    
    if request.method == 'POST':
        quantite = int(request.POST['quantite'])
        qtite_plaq = int(request.POST['quantite_plaq'])
        travailleur_id = int(request.POST['travailleur'])

        if travailleur_id:
            travailleur = Travailleurs.objects.get(id=travailleur_id)
            if quantite <= medicament.quantité_stock and qtite_plaq <= medicament.quantité_detail :
                medicament.quantité_stock -= quantite
                medicament.quantité_detail -= qtite_plaq
                medicament.save()
                
                transaction = Transactions(medicaments=medicament, type_transaction=Transactions.VENTE, quantite=quantite, quantite_plaq=qtite_plaq, travailleurs=travailleur, user=request.user)
                transaction.save(request=request) # Passer l'utilisateur connecté à la transaction
                
                return redirect('stocks:list_transaction')
            else:
                # Gérer le cas où la quantité demandée est supérieure à la quantité disponible
                raise ValueError("La quantité demandée dépasse la quantité disponible.")
        else:
            messages.error(request, "Veuillez sélectionner un fournisseur.")
        
    return render(request, 'transaction/retirer_medoc.html', {'medicament': medicament, 'travailleurs':travailleurs})

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def exporter_liste_medicaments(request):
    medocs = Medicaments.objects.all()

    # Créer un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active

    # Ajouter les en-têtes des colonnes
    sheet['A1'] = 'ID'
    # sheet['B1'] = 'Date de création'
    sheet['B1'] = 'Code'
    sheet['C1'] = 'Nom du médicament'
    sheet['D1'] = 'Nom commercial'
    sheet['E1'] = 'Quantités en boite'
    sheet['F1'] = 'Unité'

    # Remplir les données des médicaments
    row = 2
    for medoc in medocs:
        sheet.cell(row=row, column=1).value = medoc.id
        # Convertir la date de création en un objet datetime sans fuseau horaire
        #car excell ne prend pas en charge les fuseaux horaires
        # date_creation = medoc.date_creation.replace(tzinfo=None)
        # sheet.cell(row=row, column=2).value = date_creation
        # sheet.cell(row=row, column=2).value = medoc.date_creation
        sheet.cell(row=row, column=2).value = medoc.code_medoc
        sheet.cell(row=row, column=3).value = medoc.nom_medoc
        sheet.cell(row=row, column=4).value = medoc.nom_commercial
        sheet.cell(row=row, column=5).value = medoc.quantité_stock
        sheet.cell(row=row, column=6).value = medoc.unité
        row += 1

    # Configurer la réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Stock_Actuel.xlsx"'

    # Enregistrer le classeur Excel dans la réponse HTTP
    workbook.save(response)
  
    return response


class HomeWorkView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Travailleurs
    template_name = "users/worker.html"
    context_object_name = "travailleur"

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        travailleur = Travailleurs.objects.all()
        paginator = Paginator(travailleur, 10) #chaque page affiche 10Travailleurs
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['travailleur'] = page_obj
        return context
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        nom = request.POST.get('nom')
        matricule = request.POST.get('matricule')

        travailleur = Travailleurs.objects.all()
    
        if nom:
            travailleur = Travailleurs.search_by_name(nom)
        if matricule:
            travailleur = Travailleurs.search_by_matricule(matricule)

        context = {
            'travailleur': travailleur
        }

        return render(request, 'users/worker.html', context)


class HomeFournView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Fournisseurs
    template_name = "users/fournisseur.html"
    context_object_name = "fournisseur"

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()
    
    def get_test_func(self):
        return self.test_func

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fournisseur = Fournisseurs.objects.all()
        paginator = Paginator(fournisseur, 10) #chaque page affiche 10 fournisseurs
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['fournisseur'] = page_obj
        return context
 
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()

        if request.method == 'POST':
            return self.post(request, *args, **kwargs)
        else:
            return self.get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # Gérer la logique pour les requêtes POST ici
        nom = request.POST.get('nom')
        matricule = request.POST.get('matricule')

        fournisseur = Fournisseurs.objects.all()
    
        if nom:
            fournisseur = Fournisseurs.search_by_name(nom)
        if matricule:
            fournisseur = Travailleurs.search_by_matricule(matricule)

        context = {
            'fournisseur': fournisseur
        }

        return render(request, 'users/fournisseur.html', context)


class CreateWorkerView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Travailleurs
    template_name = "users/worker_create.html"
    form_class = TravailleursForm
    success_url = reverse_lazy('stocks:list_worker')

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()

    def form_invalid(self, form):
        # Renvoie les erreurs de validation sous forme de réponse JSON
        return JsonResponse({'errors': form.errors})
    
    def form_valid(self, form):
        # Enregistrer le formulaire si valide
        self.object = form.save()
        # Renvoyer une réponse JSON avec succès
        return JsonResponse({'success': True}) 

class CreateFournView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Fournisseurs
    template_name = "users/fournisseur_create.html"
    form_class = FournisseursForm
    success_url = reverse_lazy('stocks:list_fournisseur')

    def test_func(self):
        return self.request.user.groups.filter(name='Admins').exists()

    def form_invalid(self, form):
        # Renvoie les erreurs de validation sous forme de réponse JSON
        return JsonResponse({'errors': form.errors})
    
    def form_valid(self, form):
        # Enregistrer le formulaire si valide
        self.object = form.save()
        # Renvoyer une réponse JSON avec succès
        return JsonResponse({'success': True}) 


@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def export_transactions(request):
    transactions = Transactions.objects.all()  # Récupérez les transactions selon vos critères de filtrage

    # Créer un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active

    # Ajouter les en-têtes des colonnes
    sheet.append(['Type de transaction', 'Quantité en boite', 'Quantité en plaquettes', 'Date de transaction', 'Médicament', 'Travailleur', 'Fournisseur'])

    # Ajouter les données des transactions
    for transaction in transactions:
        #Si le champ existe on assigne sa valeur a la variable sinon une chaine vide
        medicament = transaction.medicaments.nom_medoc if transaction.medicaments else ""
        travailleur = transaction.travailleurs.nom if transaction.travailleurs else ""
        fournisseur = transaction.fournisseur.nom if transaction.fournisseur else ""

        # date_transaction = transaction.date_transaction.replace(tzinfo=None)
        date_formatted = formats.date_format(transaction.date_transaction, "SHORT_DATETIME_FORMAT")


        sheet.append([transaction.type_transaction, transaction.quantite, transaction.quantite_plaq, date_formatted, medicament, travailleur, fournisseur])

    # Créer la réponse HTTP et définir le type de contenu
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'

    # Enregistrer le classeur Excel dans la réponse
    workbook.save(response)

    return response   

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def expoter(request):
    return render(request, 'transaction/export.html')

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def export_etat(request):
  # Récupérer les paramètres de la requête GET
    categorie = request.GET.get('categorie')
    mois = request.GET.get('mois')
    annee = request.GET.get('annee')
    date1 = request.GET.get('date1')
    date2 = request.GET.get('date2')   
    medicamen = request.GET.get('medicamen') 

    # Filtrer les transactions en fonction des paramètres
    transactions = Transactions.objects.all()
    if medicamen:
        transactions = transactions.filter(medicaments__nom_medoc =medicamen)
    if categorie:
        transactions = transactions.filter(type_transaction=categorie)
    if mois and mois.strip():
        mois = int(mois)
        transactions = transactions.filter(date_transaction__month=mois)
    if annee and annee.strip():
        annee = int(annee)
        transactions = transactions.filter(date_transaction__year=annee)
    if date1 and date2:
        transactions = transactions.filter(date_transaction__range=[date1, date2])

    # Créer un nouveau classeur Excel
    workbook = Workbook()
    sheet = workbook.active

    # Ajouter les en-têtes des colonnes
    headers = ['Type de transaction', 'Quantité en boite', 'Quantité en plaquettes', 'Date de transaction', 'Nom du médicament', 'Travailleur', 'Fournisseur']
    for col_num, header in enumerate(headers, 1): 
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Ajouter les données des transactions
    for row_num, transaction in enumerate(transactions, 2):
        date_formatted = formats.date_format(transaction.date_transaction, "SHORT_DATETIME_FORMAT")
        medicament = transaction.medicaments.nom_medoc if transaction.medicaments else ""
        travailleur = transaction.travailleurs.nom if transaction.travailleurs else ""
        fournisseur = transaction.fournisseur.nom if transaction.fournisseur else ""
        sheet[f'A{row_num}'] = transaction.type_transaction
        sheet[f'B{row_num}'] = transaction.quantite
        sheet[f'C{row_num}'] = transaction.quantite_plaq
        sheet[f'D{row_num}'] = date_formatted
        sheet[f'E{row_num}'] = medicament
        sheet[f'F{row_num}'] = travailleur
        sheet[f'G{row_num}'] = fournisseur

    # Ajouter le total des quantités en bas du rapport
    total_quantite = transactions.aggregate(Sum('quantite'))['quantite__sum']
    total_quantite_plaq = transactions.aggregate(Sum('quantite_plaq'))['quantite_plaq__sum']
    last_row = len(transactions) + 2
    sheet[f'B{last_row}'] = total_quantite
    sheet[f'C{last_row}'] = total_quantite_plaq

    # Créer la réponse HTTP et définir le type de contenu
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_transaction.xlsx'

    # Enregistrer le classeur Excel dans la réponse
    workbook.save(response)

    return response


def generate_years():
    return range(2010, 2061)

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def statistiques_details(request):
    month = request.GET.get('month')
    year = request.GET.get('year')
    societe = request.GET.get('societe')

    transactions = Transactions.objects.all()

    if month:
        transactions = transactions.filter(date_transaction__month=month)
    if year:
        transactions = transactions.filter(date_transaction__year=year)
    if societe:
        transactions = transactions.filter(travailleurs__societe=societe)

    total_sorties_boite = transactions.filter(type_transaction=Transactions.VENTE).aggregate(total=Sum('quantite'))['total']
    total_sorties_plaq = transactions.filter(type_transaction=Transactions.VENTE).aggregate(total=Sum('quantite_plaq'))['total']

    statistiques = transactions.filter(
        type_transaction=Transactions.VENTE
    ).values(
        'medicaments__nom_medoc', 'travailleurs__societe'
    ).annotate(
        sorties_total=Sum('quantite')
    ).annotate(pourcentage_sorties=(Sum('quantite') * 100.0) / total_sorties_boite
    ).order_by('-sorties_total').annotate(
        sorties_total_plaq=Sum('quantite_plaq')
    ).annotate(pourcentage=(Sum('quantite_plaq') * 100.0) / total_sorties_plaq)

    return render(request, 'other/statistiques.html', {'statistiques': statistiques, 'total_sorties_boite': total_sorties_boite, 'years': generate_years(), 'total_sorties_plaq': total_sorties_plaq})

@user_passes_test(lambda user: user.groups.filter(name='Admins').exists())
@login_required
def statistiques_global(request):
    month = request.GET.get('month')
    year = request.GET.get('year')
    societe = request.GET.get('societe')
    societe_regie = request.GET.get('societe_regie')

    transactions = Transactions.objects.all()

    if month:
        transactions = transactions.filter(date_transaction__month=month)
    if year:
        transactions = transactions.filter(date_transaction__year=year)
    if societe:
        transactions = transactions.filter(travailleurs__societe=societe)
    if societe_regie:
        transactions = transactions.filter(travailleurs__societe_regie=societe_regie)

    total_sorties_boite = transactions.filter(type_transaction=Transactions.VENTE).aggregate(total=Sum('quantite'))['total']
    total_sorties_plaq = transactions.filter(type_transaction=Transactions.VENTE).aggregate(total=Sum('quantite_plaq'))['total']

    statistiques_societe = transactions.filter(
        type_transaction=Transactions.VENTE
    ).values(
        'travailleurs__societe'
    ).annotate(
        sorties_total=Sum('quantite')
    ).annotate(
        pourcentage_sorties=(Sum('quantite') * 100.0) / total_sorties_boite
    ).order_by('-sorties_total').annotate(
        sorties_total_plaq=Sum('quantite_plaq')
    ).annotate(
        pourcentage=(Sum('quantite_plaq') * 100.0) / total_sorties_plaq
    )

    statistiques_societe_regie = transactions.filter(
        type_transaction=Transactions.VENTE, travailleurs__societe_regie__isnull=False,travailleurs__societe__exact='REGIE', #On extraire les transactions vente ou la societe exacte est regie et ou les societe regie ne sont pas null
    ).values(
        'travailleurs__societe_regie'
    ).annotate(
        sorties_total=Sum('quantite')
    ).annotate(
        pourcentage_sorties=(Sum('quantite') * 100.0) / total_sorties_boite
    ).order_by('-sorties_total').annotate(
        sorties_total_plaq=Sum('quantite_plaq')
    ).annotate(
        pourcentage=(Sum('quantite_plaq') * 100.0) / total_sorties_plaq
    )

    return render(request, 'other/stats_global.html', {
        'statistiques_societe': statistiques_societe,
        'statistiques_societe_regie': statistiques_societe_regie,
        'total_sorties_boite': total_sorties_boite,
        'years': generate_years(),
        'total_sorties_plaq': total_sorties_plaq
    })
 